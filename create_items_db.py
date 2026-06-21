from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()

HDR_BG   = 'FF1a1a2e'
HDR_FG   = 'FFFFD700'
ROW_ALT  = 'FF16213e'
ROW_NORM = 'FF0f3460'
TEXT_FG  = 'FFe0e0e0'

def hdr(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=HDR_FG, name='Arial', size=10)
    c.fill      = PatternFill('solid', start_color=HDR_BG)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def cell(ws, row, col, value=''):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(color=TEXT_FG, name='Arial', size=9)
    c.fill      = PatternFill('solid', start_color=ROW_ALT if row % 2 == 0 else ROW_NORM)
    c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 20
    return c

# ── Materiais ────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Materiais'
ws1.sheet_properties.tabColor = '888888'
ws1.row_dimensions[1].height = 30

headers1 = ['Arquivo', 'Nome', 'Mob que dropa', 'Drop chance (%)', 'Fase', 'Observacoes']
widths1  = [28, 22, 18, 16, 8, 30]
for i, (h, w) in enumerate(zip(headers1, widths1), 1):
    hdr(ws1, 1, i, h, w)

rows1 = [
    ['itm_herb.png',          'Erva da Floresta',     '', '', '1', ''],
    ['itm_goblin_coin.png',   'Moeda de Goblin',      '', '', '1', ''],
    ['itm_bone_fragment.png', 'Fragmento de Osso',    '', '', '1', ''],
    ['itm_wolf_pelt.png',     'Pele de Lobo',         '', '', '1', ''],
    ['itm_wolf_fang.png',     'Presa de Lobo',        '', '', '1', ''],
    ['itm_orc_tooth.png',     'Dente de Orc',         '', '', '1', ''],
    ['itm_shard_metal.png',   'Fragmento de Armadura','', '', '1', ''],
    ['itm_purple_gem.png',    'Gema Roxa',            '', '', '',  ''],
    ['itm_rune_coin.png',     'Moeda com Runa',       '', '', '',  ''],
]
for r, row in enumerate(rows1, 2):
    for c, val in enumerate(row, 1):
        cell(ws1, r, c, val)

# ── Consumiveis ───────────────────────────────────────────────
ws2 = wb.create_sheet('Consumiveis')
ws2.sheet_properties.tabColor = 'FF5c5c'
ws2.row_dimensions[1].height = 30

headers2 = ['Arquivo', 'Nome', 'Efeito', 'Valor do efeito', 'Duracao (s)', 'Drop chance (%)', 'Mob que dropa', 'Fase', 'Observacoes']
widths2  = [30, 22, 20, 16, 12, 16, 18, 8, 28]
for i, (h, w) in enumerate(zip(headers2, widths2), 1):
    hdr(ws2, 1, i, h, w)

rows2 = [
    ['itm_small_health_potion.png', 'Pocao Menor de Vida', 'Cura HP',  '', '-', '', '', '1', ''],
    ['itm_small_mana_potion.png',   'Pocao de Mana',       '?',        '', '-', '', '', '',  'Definir efeito'],
    ['itm_strength_elixir.png',     'Elixir de Forca',     '+ATK temp','', '',  '', '', '',  ''],
]
for r, row in enumerate(rows2, 2):
    for c, val in enumerate(row, 1):
        cell(ws2, r, c, val)

# ── Armas ─────────────────────────────────────────────────────
ws3 = wb.create_sheet('Armas')
ws3.sheet_properties.tabColor = 'F0C419'
ws3.row_dimensions[1].height = 30

headers3 = ['Arquivo', 'Nome', 'Hero(s) que pode usar', 'Raridade', 'Bonus ATK', 'Bonus HP', 'Outro bonus', 'Drop chance (%)', 'Mob que dropa', 'Fase', 'Observacoes']
widths3  = [30, 22, 22, 12, 10, 10, 18, 16, 18, 8, 28]
for i, (h, w) in enumerate(zip(headers3, widths3), 1):
    hdr(ws3, 1, i, h, w)

rows3 = [
    ['itm_small_wooden_sword.png', 'Espada de Madeira',   '', 'Comum',   '', '', '', '', '', '1', ''],
    ['itm_long_wooden_sword.png',  'Espada Longa Madeira','', 'Comum',   '', '', '', '', '', '1', ''],
    ['itm_iron_sword.png',         'Espada de Ferro',     '', 'Incomum', '', '', '', '', '', '1', ''],
    ['itm_dark_shadow_blade.png',  'Lamina das Sombras',  '', 'Epico',   '', '', '', '', '', '',  ''],
    ['itm_small_wooden_bow.png',   'Arco de Madeira',     '', 'Comum',   '', '', '', '', '', '1', 'Cacador'],
    ['itm_wooden_staff.png',       'Cajado de Madeira',   '', 'Comum',   '', '', '', '', '', '1', 'Mago/Clerigo'],
    ['itm_wooden_mace.png',        'Maca de Madeira',     '', 'Comum',   '', '', '', '', '', '1', ''],
]
for r, row in enumerate(rows3, 2):
    for c, val in enumerate(row, 1):
        cell(ws3, r, c, val)

# ── Armaduras ─────────────────────────────────────────────────
ws4 = wb.create_sheet('Armaduras')
ws4.sheet_properties.tabColor = '4A90D9'
ws4.row_dimensions[1].height = 30

headers4 = ['Arquivo', 'Nome', 'Hero(s) que pode usar', 'Raridade', 'Bonus HP', 'Outro bonus', 'Drop chance (%)', 'Mob que dropa', 'Fase', 'Observacoes']
widths4  = [30, 22, 22, 12, 10, 18, 16, 18, 8, 28]
for i, (h, w) in enumerate(zip(headers4, widths4), 1):
    hdr(ws4, 1, i, h, w)

rows4 = [
    ['itm_brown_leather_armor.png', 'Armadura de Couro', '', 'Comum',   '', '', '', '', '1', ''],
    ['itm_dark_wolf_cape.png',      'Capa de Lobo',      '', 'Incomum', '', '', '', '', '1', ''],
    ['iron_chainmail_armor.png',    'Cota de Malha',     '', 'Raro',    '', '', '', '', '1', 'Renomear: itm_chainmail_armor.png'],
]
for r, row in enumerate(rows4, 2):
    for c, val in enumerate(row, 1):
        cell(ws4, r, c, val)

# ── Acessorios ────────────────────────────────────────────────
ws5 = wb.create_sheet('Acessorios')
ws5.sheet_properties.tabColor = 'A855F7'
ws5.row_dimensions[1].height = 30

headers5 = ['Arquivo', 'Nome', 'Hero(s) que pode usar', 'Raridade', 'Bonus ATK', 'Bonus HP', 'Outro bonus', 'Drop chance (%)', 'Mob que dropa', 'Fase', 'Observacoes']
widths5  = [30, 22, 22, 12, 10, 10, 18, 16, 18, 8, 28]
for i, (h, w) in enumerate(zip(headers5, widths5), 1):
    hdr(ws5, 1, i, h, w)

rows5 = [
    ['itm_green_forest_amulet.png', 'Amuleto da Floresta', '', 'Incomum', '', '', '', '', '', '1', ''],
    ['itm_wolf_fangs_necklace.png', 'Colar de Presas',     '', 'Raro',    '', '', '', '', '', '1', ''],
    ['itm_gold_ring_purple.png',    'Anel da Alma',        '', 'Epico',   '', '', '', '', '', '',  ''],
]
for r, row in enumerate(rows5, 2):
    for c, val in enumerate(row, 1):
        cell(ws5, r, c, val)

wb.save('assets/items/items_database.xlsx')
print('Criado com sucesso!')
